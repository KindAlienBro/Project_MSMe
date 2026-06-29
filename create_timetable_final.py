"""
Complete timetable_template.xlsx generator - FINAL VERSION
Incorporates all corrections from user feedback.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# SHEET 1: Faculties
# ============================================================
ws_fac = wb.active
ws_fac.title = "Faculties"
ws_fac.append(["id", "name", "designation", "max_hours"])

faculties = [
    ("AI00355", "Dr. Vijayashekar S S", "Professor", 6),
    ("AI00378", "Dr. Kavitha Nair R", "Assoc. Prof", 15),
    ("AI00755", "Dr. S Anu Pallavi", "Assoc. Prof", 16),
    ("AI00425", "Mr. Jovin Deglus", "Asst. Prof", 16),
    ("AI00604", "Mr. Syed Musadiq Illahi", "Asst. Prof", 18),
    ("AI00893", "Mr. Nanda Kumar N", "Asst. Prof", 18),
    ("AI00822", "Ms. Surbhi", "Asst. Prof", 18),
    ("AI01146", "Ms. Vinutha M", "Asst. Prof", 18),
    ("AI00825", "Mr. Mhd Tahir Mirji", "Asst. Prof", 18),
    ("AI00917", "Mr. Sanjay P", "Asst. Prof", 18),
    ("AI01204", "Mr. Abhijith S", "Asst. Prof", 17),
    ("AI01093", "Mr. Praveen Arokia Raj", "Asst. Prof", 18),
    ("NTA1", "Nandita Tarikeri", "TA", 50),
    ("NF1", "Ranjan Kumar", "Asst. Prof", 18),
    ("NF2", "Chalasani Jayanth", "Asst. Prof", 18),
    ("NF3", "NF3", "Asst. Prof", 18),
    ("BASAV", "Basavaraj", "Instructor", 50),
    ("soft", "Soft Skills Trainer", "Trainer", 50),
    ("MBA1", "Mr. Yogesh Dixith", "Asst. Prof", 18),
    ("CIVIL", "CIVIL_1", "Asst. Prof", 18),
    ("MATHS", "Maths Department", "External", 50),
    ("oe_fac_1", "OE Teacher 1", "Guest", 18),
    ("oe_fac_2", "OE Teacher 2", "Guest", 18),
    ("oe_fac_3", "OE Teacher 3 (Sem 7)", "Guest", 18),
]
for f in faculties:
    ws_fac.append(list(f))

# ============================================================
# SHEET 2: Subjects
# ============================================================
ws_sub = wb.create_sheet("Subjects")
ws_sub.append(["code", "name", "type", "credits", "is_core", "is_heavy"])

subjects = [
    # === 7th Sem AIML ===
    ("DL&RL", "Deep Learning & RL", "THEORY", 3, True, True),
    ("DL&RL Lab", "DL&RL Lab", "LAB", 1, True, False),
    ("ML II", "Machine Learning-II", "THEORY", 3, True, True),
    ("ML II Lab", "ML-II Lab", "LAB", 1, True, False),
    ("DS&P", "Data Security & Privacy", "THEORY", 4, True, True),
    ("BA", "Business Analytics", "THEORY", 3, False, False),
    ("BDA", "Big Data Analytics", "THEORY", 3, False, False),
    ("OE (Sem 7)", "Open Elective (Sem 7)", "THEORY", 3, False, False),
    ("MP Phase-II (BAI)", "Major Project Phase-II", "LAB", 6, True, True),
    ("Intro DBMS", "Introduction to DBMS", "THEORY", 3, True, False),
    # === 7th Sem DS ===
    ("PP", "Parallel Programming", "THEORY", 3, True, True),
    ("PP Lab", "PP Lab", "LAB", 1, False, False),
    ("SML for DS", "Statistical ML for DS", "THEORY", 3, True, True),
    ("SML for DS Lab", "Statistical ML for DS Lab", "LAB", 1, False, False),
    ("CNS", "Cryptography and Network Security", "THEORY", 4, True, True),
    ("DL", "Deep Learning", "THEORY", 3, False, False),
    ("MP Phase-II (BCD)", "Major Project Phase-II", "LAB", 6, True, True),
    # === 5th Sem Common ===
    ("SEPM", "Software Engg & Project Management", "THEORY", 4, True, True),
    ("CN", "Computer Networks", "THEORY", 3, True, True),
    ("CN Lab", "CN Lab", "LAB", 1, True, False),
    ("TOC", "Theory of Computation", "THEORY", 5, True, True),
    ("DV Lab", "Data Visualization Lab", "LAB", 1, True, False),
    ("Mini Project (BAI)", "Mini Project", "LAB", 1, True, False),
    ("Mini Project (BCD)", "Mini Project (DS)", "LAB", 1, True, False),
    ("RM&IPR", "Research Methodology & IPR", "THEORY", 1, False, False),
    ("EVS", "Environmental Studies", "THEORY", 1, False, False),
    ("CV", "Computer Vision", "THEORY", 3, False, False),
    ("EDA", "Exploratory Data Analysis", "THEORY", 3, False, False),
    ("NoSQL DB", "NoSQL Databases", "THEORY", 3, True, False),
    ("DW", "Data Warehousing", "THEORY", 3, True, False),
    # === 3rd Sem ===
    ("Maths (Prob & Stats)", "Probability, Distributions and Statistics", "THEORY", 4, True, True),
    ("DDCO", "Digital Design & Computer Organization", "THEORY", 5, True, True),
    ("OS", "Operating Systems", "THEORY", 3, True, True),
    ("DSA", "Data Structures and Applications", "THEORY", 3, True, True),
    ("DSA Lab", "DSA Lab", "LAB", 1, True, True),
    ("OOP with Java", "OOP with JAVA", "THEORY", 1, False, False),
    ("OOP with Java Lab", "OOP with JAVA Lab", "LAB", 1, False, False),
    ("EDA Lab", "Exploratory Data Analysis Lab", "LAB", 1, False, False),
    ("PM with Git", "Project Management with GIT", "LAB", 1, False, False),
    ("PBL", "Project Based Learning", "LAB", 1, True, False),
    ("Maths LE", "Mathematics for Lateral Entry Students", "THEORY", 1, False, False),
    ("NCMC 3", "NCMC (Sem 3)", "THEORY", 1, False, False),
    ("NCMC 5", "NCMC (Sem 5)", "THEORY", 1, False, False),
    # === Non-academic ===
    ("NSS", "NSS", "LAB", 1, False, False),
    ("Library", "Library Hour", "THEORY", 1, False, False),
    ("Student Hr", "Student Hour", "THEORY", 1, False, False),
    ("Faculty Hr", "Faculty Hour", "THEORY", 1, False, False),
    ("Study Hr", "Study Hour", "THEORY", 1, False, False),
    ("Soft Skills", "Soft Skills", "SOFTSKILL", 1, True, False),
    ("Forum", "Forum", "FORUM", 1, True, False),
]
for s in subjects:
    ws_sub.append(list(s))

# ============================================================
# SHEET 3: Sections
# ============================================================
ws_sec = wb.create_sheet("Sections")
ws_sec.append(["id", "semester", "strength"])

sections = [
    ("7A", "7A AI", 71), ("7B", "7B AI", 73),
    ("5A", "5A AI", 61), ("5B", "5B AI", 67),
    ("3A", "3A AI", 65), ("3B", "3B AI", 65),
    ("7DS", "7A DS", 70), ("5DS", "5A DS", 70), ("3DS", "3A DS", 70),
    ("OE_AI", "OE_1", 70),
    ("7A-B1", "7A AI", 35), ("7A-B2", "7A AI", 36),
    ("7B-B1", "7B AI", 36), ("7B-B2", "7B AI", 37),
    ("7DS-B1", "7A DS", 35), ("7DS-B2", "7A DS", 35),
    ("5A-B1", "5A AI", 30), ("5A-B2", "5A AI", 31),
    ("5B-B1", "5B AI", 33), ("5B-B2", "5B AI", 34),
    ("5DS-B1", "5A DS", 35), ("5DS-B2", "5A DS", 35),
    ("3A-B1", "3A AI", 32), ("3A-B2", "3A AI", 33),
    ("3B-B1", "3B AI", 32), ("3B-B2", "3B AI", 33),
    ("3DS-B1", "3A DS", 35), ("3DS-B2", "3A DS", 35),
]
for s in sections:
    ws_sec.append(list(s))

# ============================================================
# SHEET 4: Rooms
# ============================================================
ws_room = wb.create_sheet("Rooms")
ws_room.append(["id", "capacity", "is_lab", "building"])

rooms = [
    ("AI09 3F 01", 80, False, "ECE BLOCK"),
    ("AI09 3F 03", 80, False, "ECE BLOCK"),
    ("AI09 3F 08", 80, False, "ECE BLOCK"),
    ("AI09 2F 12", 80, False, "ECE BLOCK"),
    ("AI09 2F 13", 80, False, "ECE BLOCK"),
    ("AI09 2F 14", 80, False, "ECE BLOCK"),
    ("AI09 2F 08", 80, False, "ECE BLOCK"),
    ("AI09 1F 25", 80, False, "ECE BLOCK"),
    ("AI09 1F26", 80, False, "ECE BLOCK"),
    ("AI03 2F 01", 80, True, "ASD BLOCK"),
    ("AI03 2F 02", 80, True, "ASD BLOCK"),
]
for r in rooms:
    ws_room.append(list(r))

# ============================================================
# SHEET 5: Allocations (CORRECTED)
# ============================================================
ws_alloc = wb.create_sheet("Allocations")
ws_alloc.append(["faculty_id", "subject_code", "section_id", "elective_group"])

allocations = [
    # =============================================
    # 7th Sem AIML (7A, 7B)
    # =============================================
    ("AI00355", "DL&RL", "7A", ""),
    ("AI00355", "DL&RL", "7B", ""),
    ("AI00378", "DL&RL Lab", "7A-B1", "lab_pair_7A_2"),
    ("AI00378", "DL&RL Lab", "7A-B2", "lab_pair_7A_1"),
    ("AI01093", "DL&RL Lab", "7B-B1", "lab_pair_7B_2"),
    ("AI01093", "DL&RL Lab", "7B-B2", "lab_pair_7B_1"),
    ("AI00604", "ML II", "7A", ""),
    ("AI01204", "ML II", "7B", ""),
    ("AI00604", "ML II Lab", "7A-B1", "lab_pair_7A_1"),
    ("AI00604", "ML II Lab", "7A-B2", "lab_pair_7A_2"),
    ("AI01204", "ML II Lab", "7B-B1", "lab_pair_7B_1"),
    ("AI01204", "ML II Lab", "7B-B2", "lab_pair_7B_2"),
    ("AI00425", "DS&P", "7A", ""),
    ("AI00425", "DS&P", "7B", ""),
    ("MBA1", "BA", "7A", "ELEC_7"),
    ("MBA1", "BA", "7B", "ELEC_7"),
    ("MBA1", "BA", "7DS", "ELEC_7"),
    ("AI00822", "BDA", "7A", "ELEC_7"),
    ("AI00822", "BDA", "7B", "ELEC_7"),
    ("AI00893", "DL", "7DS", "ELEC_7"),
    # FIX #1: Intro DBMS for BOTH 7A and 7B
    ("AI00825", "Intro DBMS", "7A", ""),
    ("AI00825", "Intro DBMS", "7B", ""),
    ("AI00425", "MP Phase-II (BAI)", "7A", ""),
    ("AI00755", "MP Phase-II (BAI)", "7B", ""),
    ("oe_fac_1", "OE (Sem 7)", "7A", ""),
    ("oe_fac_2", "OE (Sem 7)", "7B", ""),
    ("oe_fac_3", "OE (Sem 7)", "7DS", ""),

    # =============================================
    # 7th Sem DS (7DS)
    # =============================================
    ("NF2", "PP", "7DS", ""),
    ("NF2", "PP Lab", "7DS-B1", ""),
    ("NF2", "PP Lab", "7DS-B2", ""),
    ("AI00378", "SML for DS", "7DS", ""),
    ("AI00378", "SML for DS Lab", "7DS-B1", ""),
    ("AI00378", "SML for DS Lab", "7DS-B2", ""),
    ("AI00425", "CNS", "7DS", ""),
    ("AI00378", "MP Phase-II (BCD)", "7DS", ""),

    # =============================================
    # 5th Sem AIML (5A, 5B)
    # =============================================
    ("NF3", "SEPM", "5A", ""),
    ("NF3", "SEPM", "5B", ""),
    ("AI00755", "CN", "5A", ""),
    ("AI00755", "CN", "5B", ""),
    ("AI00755", "CN Lab", "5A-B1", ""),
    ("AI00755", "CN Lab", "5A-B2", ""),
    ("AI00755", "CN Lab", "5B-B1", ""),
    ("AI00755", "CN Lab", "5B-B2", ""),
    ("AI01146", "TOC", "5A", ""),
    ("AI01146", "TOC", "5B", ""),
    # FIX #5: DV Lab per work allocation v5
    ("AI00755", "DV Lab", "5A-B1", ""),     # Dr. Anu Pallavi - 5A B1
    ("AI01204", "DV Lab", "5A-B2", ""),     # Mr. Abhijith - 5A B2 (was Jovin, fixed)
    ("AI00425", "DV Lab", "5B-B1", ""),     # Mr. Jovin - 5B B1
    ("AI00425", "DV Lab", "5B-B2", ""),     # Mr. Jovin - 5B B2 (was NF1, fixed)
    ("AI00604", "CV", "5A", "ELEC_5_AIML"),
    ("AI00822", "EDA", "5A", "ELEC_5_AIML"),
    ("AI00604", "CV", "5B", "ELEC_5_AIML"),
    ("AI00822", "EDA", "5B", "ELEC_5_AIML"),
    ("AI00604", "Mini Project (BAI)", "5A", ""),
    ("AI01093", "Mini Project (BAI)", "5B", ""),
    ("BASAV", "RM&IPR", "5A", ""),
    ("BASAV", "RM&IPR", "5B", ""),
    ("CIVIL", "EVS", "5A", ""),
    ("CIVIL", "EVS", "5B", ""),
    # FIX: Add NCMC 5th sem AIML
    ("AI00893", "NCMC 5", "5A", ""),
    ("AI00893", "NCMC 5", "5B", ""),

    # =============================================
    # 5th Sem DS (5DS)
    # =============================================
    ("NF1", "SEPM", "5DS", ""),
    ("AI00893", "CN", "5DS", ""),
    ("AI00893", "CN Lab", "5DS-B1", ""),
    ("AI00893", "CN Lab", "5DS-B2", ""),
    ("AI01146", "TOC", "5DS", ""),
    ("AI00604", "DV Lab", "5DS-B1", ""),
    ("AI01093", "DV Lab", "5DS-B2", ""),
    ("AI00825", "NoSQL DB", "5DS", "ELEC_5_DS"),
    ("AI01204", "DW", "5DS", "ELEC_5_DS"),
    # FIX #6: Mini Project DS - both Kavitha (primary) and Tahir (co-guide)
    ("AI00378", "Mini Project (BCD)", "5DS", ""),
    ("AI00825", "Mini Project (BCD)", "5DS", ""),
    ("BASAV", "RM&IPR", "5DS", ""),
    ("CIVIL", "EVS", "5DS", ""),
    # FIX: NCMC 5th sem DS
    ("AI00917", "NCMC 5", "5DS", ""),

    # =============================================
    # 3rd Sem AIML (3A, 3B)
    # =============================================
    # FIX #4: Add Maths
    ("MATHS", "Maths (Prob & Stats)", "3A", ""),
    ("MATHS", "Maths (Prob & Stats)", "3B", ""),
    ("NF1", "DDCO", "3A", ""),
    ("NF1", "DDCO", "3B", ""),
    ("AI00917", "OS", "3A", ""),
    ("AI01146", "OS", "3B", ""),
    ("AI00822", "DSA", "3A", ""),
    ("AI00825", "DSA", "3B", ""),
    ("AI00822", "DSA Lab", "3A-B1", ""),
    ("AI00893", "DSA Lab", "3A-B2", ""),
    ("AI00825", "DSA Lab", "3B-B1", ""),
    ("AI00822", "DSA Lab", "3B-B2", ""),
    ("NTA1", "OOP with Java", "3A", ""),
    ("NTA1", "OOP with Java", "3B", ""),
    ("AI00378", "EDA Lab", "3A", ""),
    ("AI00378", "EDA Lab", "3B", ""),
    ("AI00917", "PM with Git", "3A", ""),
    ("AI00917", "PM with Git", "3B", ""),
    # FIX #3: PBL
    ("AI00425", "PBL", "3A", ""),
    ("AI00755", "PBL", "3A", ""),
    ("AI00425", "PBL", "3B", ""),
    ("AI00755", "PBL", "3B", ""),
    # FIX #2: NCMC
    ("AI01146", "NCMC 3", "3A", ""),
    ("AI01146", "NCMC 3", "3B", ""),
    # Maths LE
    ("MATHS", "Maths LE", "3A", ""),
    ("MATHS", "Maths LE", "3B", ""),

    # =============================================
    # 3rd Sem DS (3DS)
    # =============================================
    ("MATHS", "Maths (Prob & Stats)", "3DS", ""),
    ("NF2", "DDCO", "3DS", ""),
    ("AI00917", "OS", "3DS", ""),
    ("AI00825", "DSA", "3DS", ""),
    ("AI00893", "DSA Lab", "3DS-B1", ""),
    ("AI00917", "DSA Lab", "3DS-B2", ""),
    ("NTA1", "OOP with Java", "3DS", ""),
    ("AI00378", "EDA Lab", "3DS", ""),
    ("NF3", "PM with Git", "3DS", ""),
    # PBL DS
    ("AI00825", "PBL", "3DS", ""),
    ("AI00378", "PBL", "3DS", ""),
    # NCMC DS
    ("AI01204", "NCMC 3", "3DS", ""),
    # Maths LE DS
    ("MATHS", "Maths LE", "3DS", ""),

    # =============================================
    # Soft Skills (all sections)
    # =============================================
    ("soft", "Soft Skills", "7A", ""),
    ("soft", "Soft Skills", "7B", ""),
    ("soft", "Soft Skills", "7DS", ""),
    ("soft", "Soft Skills", "5A", ""),
    ("soft", "Soft Skills", "5B", ""),
    ("soft", "Soft Skills", "5DS", ""),
    ("soft", "Soft Skills", "3A", ""),
    ("soft", "Soft Skills", "3B", ""),
    ("soft", "Soft Skills", "3DS", ""),

    # =============================================
    # FORUM - Class Teacher (non-struck-out from photo)
    # =============================================
    ("AI00822", "Forum", "7A", ""),   # Ms. Surbhi
    ("NF3", "Forum", "7B", ""),       # NEW FACULTY
    ("AI00917", "Forum", "7DS", ""),   # Mr. Sanjay P
    ("AI00893", "Forum", "5A", ""),    # Mr. Nanda Kumar N
    ("AI01204", "Forum", "5B", ""),    # Mr. Abhijith
    ("AI01093", "Forum", "5DS", ""),   # Mr. Praveen
    ("AI00755", "Forum", "3A", ""),    # Dr. Anu Pallavi
    ("AI01146", "Forum", "3B", ""),    # Ms. Vinutha M
    ("AI00378", "Forum", "3DS", ""),   # Dr. Kavitha Nair R
]

for a in allocations:
    ws_alloc.append(list(a))

# ============================================================
# SHEET 6: Scheduling Rules
# ============================================================
ws_rules = wb.create_sheet("Scheduling Rules")
ws_rules.append(["rule_type", "subject_codes", "subject_types", "period", "max_period", "days"])
rules = [
    ("BEFORE_TIME", "", "SOFTSKILL", "", "Period 5", ""),
    ("FIXED_PERIOD", "OE (Sem 7)", "", "Period 3", "", ""),
    ("FIXED_DAYS", "OE (Sem 7)", "", "", "", "MON, TUE, WED"),
    ("BEFORE_TIME", "", "FORUM", "", "Period 5", ""),
]
for r in rules:
    ws_rules.append(list(r))

# ============================================================
# FORMATTING
# ============================================================
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

for ws in [ws_fac, ws_sub, ws_sec, ws_room, ws_alloc, ws_rules]:
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)
    ws.freeze_panes = "A2"

# ============================================================
# Save
# ============================================================
output = r"c:\Users\kinda\OneDrive\Desktop\Vorniity\Project MSMe\timetable_template_final.xlsx"
wb.save(output)
print(f"SUCCESS: {output}")
print(f"  Faculties:      {len(faculties)}")
print(f"  Subjects:       {len(subjects)}")
print(f"  Sections:       {len(sections)}")
print(f"  Rooms:          {len(rooms)}")
print(f"  Allocations:    {len(allocations)}")
print(f"  Schedule Rules: {len(rules)}")
