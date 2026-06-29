"""
Update the timetable_template.xlsx based on user's answers:
1. Add Intro DBMS for 7B (both sections)
2. Add NCMC allocations
3. Add Project Based Learning (PBL)
4. Add Maths/Probability as external faculty
5. Fix DV Lab batches per work allocation v5
6. Fix Mini Project DS (both Kavitha + Tahir)
"""

import openpyxl

wb = openpyxl.load_workbook(r'timetable_template.xlsx')

# ============================================================
# Update Faculties - Add MATHS external faculty
# ============================================================
ws_fac = wb['Faculties']
ws_fac.append(["MATHS", "Maths Department", "External", 50])

# ============================================================
# Update Subjects - Add missing subjects
# ============================================================
ws_sub = wb['Subjects']
# Check existing subject codes
existing_subjects = set()
for row in ws_sub.iter_rows(min_row=2, max_row=ws_sub.max_row, values_only=True):
    existing_subjects.add(row[0])

new_subjects = [
    ("Maths (Prob & Stats)", "Probability, Distributions and Statistics", "THEORY", 4, True, True),
    ("Maths LE", "Mathematics for Lateral Entry Students", "THEORY", 1, False, False),
    ("PBL", "Project Based Learning", "LAB", 1, True, False),
    ("NCMC 3", "NCMC (Sem 3)", "THEORY", 1, False, False),
    ("NCMC 5", "NCMC (Sem 5)", "THEORY", 1, False, False),
]

for s in new_subjects:
    if s[0] not in existing_subjects:
        ws_sub.append(list(s))
        print(f"  Added subject: {s[0]} - {s[1]}")
    else:
        print(f"  Subject already exists: {s[0]}")

# ============================================================
# Update Sections - Add 3rd sem lab batches if missing
# ============================================================
ws_sec = wb['Sections']
existing_sections = set()
for row in ws_sec.iter_rows(min_row=2, max_row=ws_sec.max_row, values_only=True):
    existing_sections.add(row[0])

# 3A batches for DSA lab, EDA lab etc.
new_sections = [
    ("3A-B1", "3A AI", 32),
    ("3A-B2", "3A AI", 33),
    ("3B-B1", "3B AI", 32),
    ("3B-B2", "3B AI", 33),
    ("3DS-B1", "3A DS", 35),
    ("3DS-B2", "3A DS", 35),
]
for s in new_sections:
    if s[0] not in existing_sections:
        ws_sec.append(list(s))
        print(f"  Added section: {s[0]}")

# ============================================================
# Update Allocations
# ============================================================
ws_alloc = wb['Allocations']

# Collect existing allocations to avoid duplicates
existing_allocs = set()
for row in ws_alloc.iter_rows(min_row=2, max_row=ws_alloc.max_row, values_only=True):
    key = (row[0], row[1], row[2])
    existing_allocs.add(key)

def add_alloc(fac_id, subject, section, elective_group=""):
    key = (fac_id, subject, section)
    if key not in existing_allocs:
        ws_alloc.append([fac_id, subject, section, elective_group])
        existing_allocs.add(key)
        print(f"  + Added: {fac_id} -> {subject} -> {section} [{elective_group}]")
    else:
        print(f"  = Already exists: {fac_id} -> {subject} -> {section}")

def remove_alloc(fac_id, subject, section):
    """Remove a specific allocation row"""
    for row_idx in range(ws_alloc.max_row, 1, -1):
        row = ws_alloc[row_idx]
        if row[0].value == fac_id and row[1].value == subject and row[2].value == section:
            ws_alloc.delete_rows(row_idx)
            existing_allocs.discard((fac_id, subject, section))
            print(f"  - Removed: {fac_id} -> {subject} -> {section}")
            return True
    return False

print("\n=== 1. Intro DBMS for 7B ===")
add_alloc("AI00825", "Intro DBMS", "7B")

print("\n=== 2. NCMC Allocations ===")
add_alloc("AI01146", "NCMC 3", "3A")   # Ms. Vinutha -> 3A-AIML
add_alloc("AI01146", "NCMC 3", "3B")   # Ms. Vinutha -> 3B-AIML
add_alloc("AI01204", "NCMC 3", "3DS")  # Mr. Abhijith -> 3A-DS
add_alloc("AI00893", "NCMC 5", "5A")   # Mr. Nandakumar -> 5A-AIML
add_alloc("AI00893", "NCMC 5", "5B")   # Mr. Nandakumar -> 5B-AIML
add_alloc("AI00917", "NCMC 5", "5DS")  # Mr. Sanjay P -> 5A-DS

print("\n=== 3. Project Based Learning ===")
add_alloc("AI00425", "PBL", "3A")   # Mr. Jovin -> 3A-AIML
add_alloc("AI00755", "PBL", "3A")   # Dr. Anu Pallavi -> 3A-AIML (co-guide)
add_alloc("AI00425", "PBL", "3B")   # Mr. Jovin -> 3B-AIML
add_alloc("AI00755", "PBL", "3B")   # Dr. Anu Pallavi -> 3B-AIML (co-guide)
add_alloc("AI00825", "PBL", "3DS")  # Mr. Tahir -> 3DS
add_alloc("AI00378", "PBL", "3DS")  # Dr. Kavitha -> 3DS (co-guide)

print("\n=== 4. Maths / Probability ===")
add_alloc("MATHS", "Maths (Prob & Stats)", "3A")
add_alloc("MATHS", "Maths (Prob & Stats)", "3B")
add_alloc("MATHS", "Maths (Prob & Stats)", "3DS")
add_alloc("MATHS", "Maths LE", "3A")
add_alloc("MATHS", "Maths LE", "3B")
add_alloc("MATHS", "Maths LE", "3DS")

print("\n=== 5. Fix DV Lab batches (per work alloc v5) ===")
# Work alloc v5 for DV Lab:
# Dr. Anu Pallavi: 5A AIML 1 batch + 5&B AIML CN Lab
# Mr. Jovin: 5B AIML 2 batches
# Mr. Abhijith: 5A AIML 1 batch
# NF1 (Ranjan): 5B AIML CO 2 batches
# NF2 (Chalasani): 5A AIML CO 1 batch + 5DS CO 2 batches
# Mr. Syed: 5DS 1 batch
# Mr. Praveen: 5DS 1 batch

# Current template DV Lab:
# AI00755 (Anu Pallavi) -> 5A-B1 ✓
# AI00425 (Jovin) -> 5A-B2 ✗ should be AI01204 (Abhijith)
# AI00425 (Jovin) -> 5B-B1 ✓
# NF1 -> 5B-B2 ✗ should be AI00425 (Jovin) per v5 (Jovin has 2 batches of 5B)
# AI00604 (Syed) -> 5DS-B1 ✓
# AI01093 (Praveen) -> 5DS-B2 ✓

# Fix 5A-B2: Change from Jovin to Abhijith
remove_alloc("AI00425", "DV Lab", "5A-B2")
add_alloc("AI01204", "DV Lab", "5A-B2")

# Fix 5B-B2: Change from NF1 to Jovin (Jovin handles 2 batches of 5B per v5)
remove_alloc("NF1", "DV Lab", "5B-B2")
add_alloc("AI00425", "DV Lab", "5B-B2")

print("\n=== 6. Fix Mini Project DS ===")
# Change from only Tahir to Kavitha as primary
remove_alloc("AI00825", "Mini Project (BCD)", "5DS")
add_alloc("AI00378", "Mini Project (BCD)", "5DS")  # Dr. Kavitha Nair R (primary)
add_alloc("AI00825", "Mini Project (BCD)", "5DS")  # Mr. Tahir (co-guide) -- wait, this would be same key
# Actually add Tahir back too since both guide
# Since both need to be allocated, let's keep both
# But we can't have same (faculty, subject, section) twice
# Let's re-add Tahir
add_alloc("AI00825", "Mini Project (BCD)", "5DS")

# ============================================================
# Save
# ============================================================
output_path = r"c:\Users\kinda\OneDrive\Desktop\Vorniity\Project MSMe\timetable_template.xlsx"
wb.save(output_path)
print(f"\nSUCCESS: Updated template saved to {output_path}")

# Verify final counts
wb2 = openpyxl.load_workbook(output_path, data_only=True)
for sn in wb2.sheetnames:
    ws = wb2[sn]
    print(f"  {sn}: {ws.max_row - 1} entries")
