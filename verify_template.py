import openpyxl

# Read the work allocation v5 in detail
wb = openpyxl.load_workbook(r'work_allocation_2026-2027.xlsx', data_only=True)
ws = wb['AIML2026-06-09-v5']

print('=== WORK ALLOCATION v5 - Full Faculty Breakdown ===')
current_faculty = ''
for row_idx in range(10, 70):
    row = ws[row_idx]
    vals = {}
    for cell in row:
        if cell.value is not None:
            vals[cell.column_letter] = cell.value
    
    if 'B' in vals:
        current_faculty = vals['B']
        fac_id = vals.get('A', '')
        desig = vals.get('C', '')
        print(f'\n--- {fac_id}. {current_faculty} ({desig}) ---')
    
    # Theory
    if 'D' in vals and 'F' in vals:
        sem = vals['D']
        sub_code = vals.get('E', '')
        sub_name = vals['F']
        sections = vals.get('G', '')
        hrs = vals.get('H', '')
        total_hrs = vals.get('I', '')
        print(f'  THEORY: {sem} | {sub_code} | {sub_name} | sections={sections} | hrs/sec={hrs} | total={total_hrs}')
    
    # Lab
    if 'K' in vals and 'M' in vals:
        sem = vals['K']
        lab_code = vals.get('L', '')
        lab_name = vals['M']
        batches = vals.get('N', '')
        hrs = vals.get('O', '')
        total = vals.get('P', '')
        print(f'  LAB:    {sem} | {lab_code} | {lab_name} | batches={batches} | hrs/batch={hrs} | total={total}')

# Now check what's in template vs work allocation
print('\n\n=== ISSUES / DISCREPANCIES ===')

wb_t = openpyxl.load_workbook(r'timetable_template.xlsx', data_only=True)
ws_a = wb_t['Allocations']

# Build faculty name map from template
ws_f = wb_t['Faculties']
fac_names = {}
for row in ws_f.iter_rows(min_row=2, max_row=ws_f.max_row, values_only=True):
    fac_names[row[0]] = row[1]

# Count allocations per faculty
from collections import defaultdict
fac_hours = defaultdict(list)
for row in ws_a.iter_rows(min_row=2, max_row=ws_a.max_row, values_only=True):
    fac_id, sub, sec = row[0], row[1], row[2]
    fac_hours[fac_id].append((sub, sec))

print('\nFaculty allocation counts:')
for fid, allocs in sorted(fac_hours.items()):
    name = fac_names.get(fid, fid)
    print(f'  {fid:12s} {name:30s} -> {len(allocs)} allocations')

# Check specific issues
print('\n\n=== THINGS TO CHECK ===')

# 1. Intro DBMS - course allocation doc says both 7A and 7B for AIML, but template only has 7A
print('\n1. Intro to DBMS (7th sem):')
print('   Course alloc doc AIML: 7A=Mr. Md Tahir Mirji, 7B=Mr. Md Tahir Mirji')
print('   Work alloc v5: 7 AIML (1 section) = Mr. Mhd Tahir')
for r in ws_a.iter_rows(min_row=2, max_row=ws_a.max_row, values_only=True):
    if r[1] == 'Intro DBMS':
        print(f'   Template: {r[0]} -> {r[1]} -> {r[2]}')

# 2. Check DS 7th sem - Intro DBMS
print('\n2. DS 7th sem Intro DBMS:')
print('   Course alloc doc DS: 7A=Mr. Md Tahir Mirji')
print('   Work alloc v5: Not present separately (same subject)')

# 3. DV Lab 5A - doc says "Dr. S Anu Pallavi +Mr. Jovin Deglus/ Mr. Abhijith S+NF2"
# Work alloc v5: Dr. Anu Pallavi 5A AIML BAIL504 (1 batch), Mr. Jovin 5B AIML (2 batches), NF1 5B AIML CO (2 batch)
print('\n3. DV Lab 5A AIML batches:')
print('   Course alloc doc: Dr. S Anu Pallavi +Mr. Jovin Deglus/ Mr. Abhijith S+NF2')
print('   Work alloc v5:')
print('     Dr. Anu Pallavi: 5A AIML DV Lab 1 batch + 5&B AIML CN Lab 4 batches')
print('     Mr. Jovin: 5B AIML DV Lab 2 batches')
print('     Mr. Abhijith: 5A AIML DV Lab 1 batch')
print('     NF1 (Ranjan): 5B AIML DV Lab CO 2 batches')
for r in ws_a.iter_rows(min_row=2, max_row=ws_a.max_row, values_only=True):
    if r[1] == 'DV Lab':
        print(f'   Template: {r[0]} ({fac_names.get(r[0], "?")}) -> {r[1]} -> {r[2]}')

# 4. NCMC is in doc but not in template
print('\n4. NCMC subjects:')
print('   3A-AIML NCMC: Ms. Vinutha M')
print('   3B-AIML NCMC: Ms. Vinutha M')
print('   5A-AIML NCMC: Mr. Nandakumar N')
print('   5B-AIML NCMC: Mr. Nandakumar N')
print('   3DS NCMC: Mr. Abhijith S')
print('   5DS NCMC: Mr. Sanjay P')
ncmc_in_template = False
for r in ws_a.iter_rows(min_row=2, max_row=ws_a.max_row, values_only=True):
    if 'NCMC' in str(r[1]):
        ncmc_in_template = True
        print(f'   Template: {r[0]} -> {r[1]} -> {r[2]}')
if not ncmc_in_template:
    print('   ** NOT in template (NCMC not allocated) **')

# 5. Project Based Learning - in doc but not in template
print('\n5. Project Based Learning (3rd sem):')
print('   3A-AIML: Mr. Jovin Deglus/Dr. S Anu Pallavi')
print('   3B-AIML: Mr. Jovin Deglus/Dr. S Anu Pallavi')
print('   3DS: Mr. Md Tahir Mirji/ Dr. Kavitha Nair R')
pbl_in_template = False
for r in ws_a.iter_rows(min_row=2, max_row=ws_a.max_row, values_only=True):
    if 'Project based' in str(r[1]) or 'PBL' in str(r[1]):
        pbl_in_template = True
        print(f'   Template: {r[0]} -> {r[1]} -> {r[2]}')
if not pbl_in_template:
    print('   ** NOT in template **')

# 6. Maths/Probability - in doc but not in template
print('\n6. Probability,Distributions and Statistics (Maths):')
print('   In course allocation docs for all 3rd sem sections')
maths_in = False
for r in ws_a.iter_rows(min_row=2, max_row=ws_a.max_row, values_only=True):
    if 'Maths' in str(r[1]) or 'Probability' in str(r[1]):
        maths_in = True
        print(f'   Template: {r[0]} -> {r[1]} -> {r[2]}')
if not maths_in:
    print('   ** NOT allocated in template (external Maths dept) **')

# 7. Maths for Lateral Entry
print('\n7. Maths for Lateral Entry:')
print('   In course allocation docs for all 3rd sem sections')
for r in ws_a.iter_rows(min_row=2, max_row=ws_a.max_row, values_only=True):
    if 'Lateral' in str(r[1]) or 'Maths LE' in str(r[1]):
        print(f'   Template: {r[0]} -> {r[1]} -> {r[2]}')

# 8. Check Intro DBMS for 7B
print('\n8. Intro to DBMS - Missing 7B?')
print('   AIML doc says both 7A and 7B get Intro DBMS from Mr. Md Tahir Mirji')
print('   BUT work allocation v5 shows only 1 section for Intro DBMS')

# 9. Mini Project DS
print('\n9. Mini Project DS:')
print('   DS doc: Dr. Kavitha Nair R/Mr. Md Tahir Mirji')
for r in ws_a.iter_rows(min_row=2, max_row=ws_a.max_row, values_only=True):
    if 'Mini Project' in str(r[1]) and 'DS' in str(r[2]):
        print(f'   Template: {r[0]} ({fac_names.get(r[0], "?")}) -> {r[1]} -> {r[2]}')

# 10. DV Lab 5DS batches check
print('\n10. DV Lab 5DS:')
print('   DS doc: Mr. Syed M Illahi+NF2/Mr. Praveen Arokiaraj+NF2')
print('   Work alloc v5: Mr. Syed 5DS 1 batch, NF2 (Chalasani) 5DS CO 2 batches, Mr. Praveen 5DS 1 batch')
for r in ws_a.iter_rows(min_row=2, max_row=ws_a.max_row, values_only=True):
    if r[1] == 'DV Lab' and 'DS' in str(r[2]):
        print(f'   Template: {r[0]} ({fac_names.get(r[0], "?")}) -> {r[1]} -> {r[2]}')
